from dataclasses import dataclass
import pandas as pd
import re
from openpyxl import load_workbook
from const import FILENAME

MESI = [
    "Gennaio",
    "Febbraio",
    "Marzo",
    "Aprile",
    "Maggio",
    "Giugno",
    "Luglio",
    "Agosto",
    "Settembre",
    "Ottobre",
    "Novembre",
    "Dicembre"
]


@dataclass
class Field:
  value: any
  length: int
  
  def __value__(self):
    return str(self.value).strip().rjust(self.length, "0")

# get a row from df by its index and return only column that are populated
def epurateNaNOfRowByIndex(df: pd.DataFrame,index: int):
  return df.iloc[index][~df.iloc[index].isna()]

# return a df without rows that haven't an 'n. buono' and datetime type as 'data'
def dropRowsWithoutIDAndDate(df: pd.DataFrame) -> pd.DataFrame:
  mask = pd.to_datetime(df['data'], errors='coerce').notna()
  return df.dropna(subset=['n. buono'])[mask]

# return a df without rows that are not Megagest
def dropRowsNotMegagest(df: pd.DataFrame) -> pd.DataFrame:
  regex = r'MEGAGEST*'
  mask = df.apply(lambda row: not any (row.astype(str).str.contains(regex, regex=True)), axis=1)
  return df[~mask]

def initDataFrame(df: pd.DateOffset)->pd.DataFrame:
  new_df = dropRowsWithoutIDAndDate(df)
  new_df = dropRowsNotMegagest(new_df)
  return new_df

def splitDecimalWithPadding(value: str, nPaddingInt: int, nPaddingDecimal : int) -> str:
  split = str(value).split('.')
  intero  = split[0].rjust(nPaddingInt, "0")
  
  decimal = ""
  if len(split) == 2 :
    decimal = split[1]
    
  decimal = decimal.rjust(nPaddingDecimal, "0")

  return intero+decimal

def fetchDataFromVenduto(filename, sheet):
  df = pd.read_excel(io=filename, sheet_name=sheet)
  df = initDataFrame(df)
  return df

def fetchMesiDalVenduto(filename):
    # Load the Excel file
    workbook = load_workbook(filename=filename, read_only=True)
    
    # Get the sheet names
    sheet_names = workbook.sheetnames
    
    # Close the workbook
    workbook.close()
    
    
    mesi_disponibili = []
    for mese in MESI:
      for sheet in sheet_names:
        # print(mese, sheet)
        if re.search(mese, sheet, re.IGNORECASE):
          mesi_disponibili.append(sheet)
    
    return mesi_disponibili
  
if __name__ == "__main__":
  print(fetchMesiDalVenduto("C:/Users/e.cavallo/Documents/fatturazione/venduto 2024.xlsx"))